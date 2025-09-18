import pandas as pd
import os
import re
from datetime import datetime
import openpyxl
from pathlib import Path

def extract_po_data_from_file(file_path):
    """Extract PO data from a single Excel file"""
    try:
        # Read the Excel file
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Get the first sheet (usually 'Quotation')
        sheet = workbook.active
        
        # Initialize data dictionary
        po_data = {
            'PO_Number': '',
            'Date': '',
            'Vendor_Name': '',
            'Description': '',
            'Quantity': ''
        }
        
        # Extract PO Number - usually in cells around B1:E4
        for row in range(1, 6):
            for col in range(1, 10):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    # Look for PO number patterns
                    if 'PO' in cell_value.upper() or 'Purchase Order' in cell_value:
                        # Check adjacent cells for the actual PO number
                        for offset in [-1, 0, 1]:
                            try:
                                adjacent_cell = sheet.cell(row=row, column=col+offset).value
                                if adjacent_cell and isinstance(adjacent_cell, str):
                                    # Match PO number patterns like 021220-01, 011724-01, etc.
                                    po_match = re.search(r'\d{6}-\d{2}', adjacent_cell)
                                    if po_match:
                                        po_data['PO_Number'] = po_match.group()
                                        break
                            except:
                                continue
        
        # Extract Date - usually near PO number
        for row in range(1, 6):
            for col in range(1, 10):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value:
                    # Check if it's a date
                    if isinstance(cell_value, datetime):
                        po_data['Date'] = cell_value.strftime('%Y-%m-%d')
                        break
                    elif isinstance(cell_value, str) and 'DATE' in cell_value.upper():
                        # Check adjacent cells for date
                        for offset in [-1, 0, 1]:
                            try:
                                adjacent_cell = sheet.cell(row=row, column=col+offset).value
                                if isinstance(adjacent_cell, datetime):
                                    po_data['Date'] = adjacent_cell.strftime('%Y-%m-%d')
                                    break
                                elif isinstance(adjacent_cell, str):
                                    # Try to parse date string
                                    date_match = re.search(r'\d{4}-\d{2}-\d{2}', adjacent_cell)
                                    if date_match:
                                        po_data['Date'] = date_match.group()
                                        break
                            except:
                                continue
            if po_data['Date']:
                break
        
        # Extract Vendor Name - usually in rows 8-15
        vendor_found = False
        for row in range(8, 16):
            if vendor_found:
                break
            for col in range(1, 5):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and len(cell_value.strip()) > 3:
                    # Skip common headers and company info
                    skip_terms = ['ADVANCED MACHINE', 'Chicago IL', 'PHONE', 'FAX', 'W Belmont', 'ATTN']
                    if not any(term in cell_value.upper() for term in skip_terms):
                        # This is likely the vendor name
                        po_data['Vendor_Name'] = cell_value.strip()
                        vendor_found = True
                        break
        
        # Extract Description and Quantity - usually in the middle section
        descriptions = []
        quantities = []
        
        # Look for description and quantity in a broader range
        description_started = False
        for row in range(10, 40):
            row_data = []
            for col in range(1, 10):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    row_data.append(cell_value.strip())
            
            # Join row data and look for meaningful content
            row_text = ' '.join(row_data).strip()
            
            # Skip empty rows
            if not row_text:
                continue
            
            # Skip header-like content
            skip_terms = ['ADVANCED MACHINE', 'Chicago IL', 'PHONE', 'FAX', 'W Belmont', 'ATTN', 
                         'Purchase Order', 'DATE', 'Customer ID', 'Cell:', 'Please call Ben']
            if any(term in row_text for term in skip_terms):
                continue
            
            # Check if this row contains "Description" header
            if 'DESCRIPTION' in row_text.upper():
                description_started = True
                continue
            
            # Look for quantities first
            qty_match = re.search(r'(\d+)\s*(ea|pcs|pieces)', row_text, re.IGNORECASE)
            if qty_match:
                quantities.append(f"{qty_match.group(1)} {qty_match.group(2)}")
            
            # Look for TOTAL row
            if 'TOTAL' in row_text.upper():
                total_match = re.search(r'(\d+)\s*(ea|pcs|pieces)', row_text, re.IGNORECASE)
                if total_match:
                    quantities.append(f"{total_match.group(1)} {total_match.group(2)}")
                continue
            
            # Capture description content - be much more inclusive
            # Include any row that contains meaningful technical content
            meaningful_content = False
            
            # Check for part numbers
            if re.search(r'P/N|PN:|PART|#\d+|\d{5,}', row_text, re.IGNORECASE):
                meaningful_content = True
            
            # Check for material specifications
            if re.search(r'MATERIAL|STEEL|ALUMINUM|BRASS|CHROME|STAINLESS|SS|RC|HRC', row_text, re.IGNORECASE):
                meaningful_content = True
            
            # Check for processes
            if re.search(r'GRINDING|PLATING|WELDING|MACHINING|TURNING|DRILLING|BORING|FINISH|COATING|ANODIZE|CHROME|HEAT|TREAT', row_text, re.IGNORECASE):
                meaningful_content = True
            
            # Check for certifications
            if re.search(r'CERT|CERTIFICATE|CERTIFICATION|REQUIRED|MIL|STD|SPEC', row_text, re.IGNORECASE):
                meaningful_content = True
            
            # Check for dimensions and tolerances
            if re.search(r'\.\d{3,}|±|\+/-|DIA|DIAMETER|OD|ID|LENGTH|THICK|DIMENSION', row_text, re.IGNORECASE):
                meaningful_content = True
            
            # Check for common manufacturing terms
            if re.search(r'VALVE|PIN|SHAFT|BRACKET|PLATE|BUSHING|SLEEVE|RING|BOLT|NUT|SCREW|WASHER|SPACER|GUIDE|PISTON|ROLLER|BEARING', row_text, re.IGNORECASE):
                meaningful_content = True
            
            # Check for process instructions
            if re.search(r'NEED|GRIND|MACHINE|DRILL|BORE|TURN|MILL|CUT|WELD|PLATE|COAT|TREAT|HARDEN|TEMPER', row_text, re.IGNORECASE):
                meaningful_content = True
            
            # Check for special notes
            if re.search(r'NOTE|SPECIAL|RUSH|ASAP|CRITICAL|MAX|MIN|TOLERANCE|DIMENSION', row_text, re.IGNORECASE):
                meaningful_content = True
            
            # If we found meaningful content, add it to descriptions
            if meaningful_content:
                # Clean up the text - remove excessive repetition
                cleaned_text = re.sub(r'(\b\w+\b)(\s+\1){2,}', r'\1', row_text)  # Remove word repetitions
                if cleaned_text and len(cleaned_text.strip()) > 3:
                    descriptions.append(cleaned_text.strip())
        
        # Combine descriptions and get the most relevant quantity
        po_data['Description'] = ' | '.join(descriptions) if descriptions else ''
        po_data['Quantity'] = quantities[-1] if quantities else ''  # Use the last/total quantity
        
        # If we didn't find quantity in the main section, try the filename
        if not po_data['Quantity']:
            filename = os.path.basename(file_path)
            qty_match = re.search(r'(\d+)\s*(ea|pcs)', filename, re.IGNORECASE)
            if qty_match:
                po_data['Quantity'] = f"{qty_match.group(1)} {qty_match.group(2)}"
        
        workbook.close()
        return po_data
        
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return None

def process_all_po_files(directory_path):
    """Process all Excel files in the Internal POs directory"""
    
    all_po_data = []
    processed_count = 0
    error_count = 0
    
    # Get all Excel files
    excel_files = []
    for file in os.listdir(directory_path):
        if file.endswith('.xlsx') and not file.startswith('~'):
            excel_files.append(os.path.join(directory_path, file))
    
    print(f"Found {len(excel_files)} Excel files to process...")
    
    # Process each file
    for file_path in excel_files:
        print(f"Processing: {os.path.basename(file_path)}")
        
        po_data = extract_po_data_from_file(file_path)
        
        if po_data:
            # Add filename for reference
            po_data['Source_File'] = os.path.basename(file_path)
            all_po_data.append(po_data)
            processed_count += 1
        else:
            error_count += 1
        
        # Progress update every 50 files
        if (processed_count + error_count) % 50 == 0:
            print(f"Progress: {processed_count + error_count}/{len(excel_files)} files processed")
    
    print(f"\nProcessing complete!")
    print(f"Successfully processed: {processed_count} files")
    print(f"Errors encountered: {error_count} files")
    
    return all_po_data

def create_consolidated_spreadsheet(po_data_list, output_filename):
    """Create a consolidated Excel spreadsheet with all PO data"""
    
    # Create DataFrame
    df = pd.DataFrame(po_data_list)
    
    # Clean and format the data
    df['PO_Number'] = df['PO_Number'].fillna('')
    df['Date'] = df['Date'].fillna('')
    df['Vendor_Name'] = df['Vendor_Name'].fillna('')
    df['Description'] = df['Description'].fillna('')
    df['Quantity'] = df['Quantity'].fillna('')
    
    # Sort by PO Number
    df = df.sort_values('PO_Number')
    
    # Reorder columns
    column_order = ['PO_Number', 'Date', 'Vendor_Name', 'Description', 'Quantity', 'Source_File']
    df = df[column_order]
    
    # Save to Excel
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Consolidated_PO_Data', index=False)
        
        # Format the worksheet
        worksheet = writer.sheets['Consolidated_PO_Data']
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 15  # PO_Number
        worksheet.column_dimensions['B'].width = 12  # Date
        worksheet.column_dimensions['C'].width = 25  # Vendor_Name
        worksheet.column_dimensions['D'].width = 60  # Description
        worksheet.column_dimensions['E'].width = 15  # Quantity
        worksheet.column_dimensions['F'].width = 40  # Source_File
        
        # Add header formatting
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
    
    print(f"Consolidated spreadsheet saved as: {output_filename}")
    print(f"Total records: {len(df)}")
    
    return df

def main():
    """Main function to run the PO data extraction"""
    
    # Set the directory path to the Dump folder
    dump_dir = "./Dump"
    
    if not os.path.exists(dump_dir):
        print(f"Error: Directory '{dump_dir}' not found!")
        return
    
    print("Starting PO data extraction process...")
    print("=" * 50)
    
    # Process all files
    po_data_list = process_all_po_files(dump_dir)
    
    if not po_data_list:
        print("No data extracted. Please check the files and try again.")
        return
    
    # Create consolidated spreadsheet
    output_filename = f"Consolidated_PO_Data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    df = create_consolidated_spreadsheet(po_data_list, output_filename)
    
    # Display summary statistics
    print("\n" + "=" * 50)
    print("EXTRACTION SUMMARY:")
    print("=" * 50)
    print(f"Total POs processed: {len(df)}")
    print(f"Unique vendors: {df['Vendor_Name'].nunique()}")
    print(f"Date range: {df['Date'].min()} to {df['Date'].max()}")
    print(f"Output file: {output_filename}")
    
    # Show sample of the data
    #print("\nSample of extracted data:")
    #print(df.head(10).to_string(index=False))

if __name__ == "__main__":
    main()
